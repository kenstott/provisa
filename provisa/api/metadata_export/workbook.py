# Copyright (c) 2026 Kenneth Stott
# Canary: 9c4d1f27-3a86-4e0b-b752-1d6f8a02c93e
#
# This source code is licensed under the Business Source License 1.1
# found in the LICENSE file in the root directory of this source tree.
#
# NOTICE: Use of this software for training artificial intelligence or
# machine learning models is strictly prohibited without explicit written
# permission from the copyright holder.

"""The model as a reviewable workbook (REQ-1592).

The config JSON is an input document written for a machine. A data steward asked to SANCTION a
model reads it in a spreadsheet — sorts it, filters it, annotates a row — and handing them JSON
makes review something only an engineer can do. This module turns the metadata-export projection
into that spreadsheet: one sheet per object type, a stable header row, a frozen header, an
autofilter, and an ``id`` column on every row so a reviewer's comment traces back to the object.

ONE WAY ONLY. The workbook is a REPORT. It is never accepted back as an edit path, and there is
deliberately no reader here: round-tripping a spreadsheet into configuration reintroduces every
ambiguity the config format exists to remove, and a stale workbook uploaded months later would
silently revert the model.

SOURCED FROM THE GOVERNED VIEW. Sheets are built from :class:`MetadataSnapshot` — the same
projection the catalog publishers serve — rather than from the raw config document, so masking,
retirement and the ``data_product`` export exclusion are already applied by the time a row is
written. Views, metrics and materialized views have no snapshot asset of their own (they are
attributes of a configured table, plus the org's metric list), so they are read from the config
NARROWED to the tables the snapshot published: a view over a table the snapshot dropped is not a
view the reviewer is being shown.
"""

# Requirements: REQ-1592

from __future__ import annotations

import io
from dataclasses import dataclass
from typing import TYPE_CHECKING, Any

from provisa.api.metadata_export.refs import metric_uri, table_ref, table_uri

if TYPE_CHECKING:
    from datetime import datetime

    from provisa.api.metadata_export.model import MetadataSnapshot
    from provisa.core.models import ProvisaConfig, Table

# The sheet names, in workbook order. Named constants because the router gates on them and the
# Report sheet lists them: a sheet identified by a literal in three places drifts in two of them.
REPORT = "Report"
SOURCES = "Sources"
TABLES = "Tables"
COLUMNS = "Columns"
RELATIONSHIPS = "Relationships"
VIEWS = "Views"
METRICS = "Metrics"
MATERIALIZED_VIEWS = "Materialized Views"
GLOSSARY = "Glossary Terms"

OBJECT_SHEETS = (
    SOURCES,
    TABLES,
    COLUMNS,
    RELATIONSHIPS,
    VIEWS,
    METRICS,
    MATERIALIZED_VIEWS,
    GLOSSARY,
)

HEADERS: dict[str, tuple[str, ...]] = {
    SOURCES: ("id", "source", "type", "tags", "description"),
    # REQ-1592: data_product is a COLUMN, not a filter — the report covers every registered table
    # and lets the reviewer filter the sheet on the flag in the spreadsheet. ``tags`` is the same
    # shape of answer for the registry tags: every taggable sheet carries one comma-delimited
    # column, so `fact`, `dimension` and `data_quality` are filtered the way `pii` is, rather
    # than each earning a column of its own.
    TABLES: ("id", "name", "aliases", "source", "domain", "data_product", "tags", "description"),
    COLUMNS: ("id", "table", "column", "aliases", "data type", "tags", "description"),
    RELATIONSHIPS: (
        "id",
        "name",
        "from table",
        "from column",
        "to table",
        "to column",
        "cardinality",
        "kind",
        "needs review",
        "version",
        "owner",
        "tags",
    ),
    VIEWS: ("id", "name", "domain", "source", "tags", "definition", "description"),
    METRICS: ("id", "metric", "expression", "datatype", "description", "ai context", "from fact"),
    MATERIALIZED_VIEWS: (
        "id",
        "name",
        "domain",
        "refresh interval (s)",
        "consistency",
        "persist",
        "incremental",
        "tags",
        "definition",
    ),
    GLOSSARY: ("id", "term", "definition", "abstract", "deprecated", "refs", "experts"),
}


@dataclass(frozen=True)
class ReportHeader:  # REQ-1592
    """What the leading Report sheet states, so a PARTIAL workbook cannot pass as a complete one."""

    org_id: str
    environment: str
    domains: str
    generated_at: "datetime"
    included: tuple[str, ...]
    omitted: tuple[str, ...]


@dataclass(frozen=True)
class _Tags:  # REQ-1592
    """The snapshot's registry tags, indexed the way the sheets address their rows.

    One comma-delimited cell per taggable row, filterable in the spreadsheet. Assigned and derived
    tags share the column deliberately: to a reviewer asking which tables are facts, `fact` is the
    same kind of answer as `pii`, and splitting them across a column per tag would make the sheet
    grow a column every time the vocabulary does. A parameterised assignment keeps its parameter
    (``entity:customer``, REQ-1467) — the parameter is what the tag says.
    """

    by_asset: dict[tuple[str, ...], str]
    by_relationship: dict[str, str]

    def asset(self, parts: tuple[str, ...]) -> str:
        return self.by_asset.get(parts, "")

    def relationship(self, edge_id: str) -> str:
        return self.by_relationship.get(edge_id, "")


def _tag_index(snapshot: "MetadataSnapshot") -> _Tags:
    by_asset: dict[tuple[str, ...], list[str]] = {}
    by_relationship: dict[str, list[str]] = {}
    for tag in snapshot.model_tags:
        if tag.asset is not None:
            bucket = by_asset.setdefault(tag.asset.parts, [])
        elif tag.relationship_id is not None:
            bucket = by_relationship.setdefault(tag.relationship_id, [])
        else:
            raise ValueError(f"model tag {tag.tag_id!r} addresses neither an asset nor an edge")
        # A tag can be assigned once and derived as well; the reviewer reads one name either way.
        if tag.tag_id not in bucket:
            bucket.append(tag.tag_id)
    return _Tags(
        by_asset={parts: ", ".join(ids) for parts, ids in by_asset.items()},
        by_relationship={edge: ", ".join(ids) for edge, ids in by_relationship.items()},
    )


def _view_definition(table: "Table") -> str:
    """The reviewable text of a view — its SQL, or the metric composition that generates it."""
    if table.view_sql is not None:
        return table.view_sql
    spec = table.view_metrics
    assert spec is not None  # only called for tables the view predicate accepted
    parts = [f"metrics: {', '.join(spec.metrics)}", f"dimensions: {', '.join(spec.dimensions)}"]
    if spec.filters:
        parts.append(f"filters: {', '.join(spec.filters)}")
    return "; ".join(parts)


def _is_view(table: "Table") -> bool:
    return table.view_sql is not None or table.view_metrics is not None


def _business_name(table: "Table") -> str:
    return table.alias or table.table_name


def _sources_rows(snapshot: "MetadataSnapshot", tags: _Tags) -> list[list[Any]]:
    return [
        [
            asset.semantic_uri,
            asset.id,
            asset.source_type,
            tags.asset(asset.ref.parts),
            asset.description,
        ]
        for asset in snapshot.sources
    ]


def _tables_rows(snapshot: "MetadataSnapshot", tags: _Tags) -> list[list[Any]]:
    return [
        [
            asset.semantic_uri,
            asset.name,
            ", ".join(asset.aliases),
            asset.source_id,
            asset.domain_id or "",
            asset.data_product,
            tags.asset(asset.ref.parts),
            asset.description,
        ]
        for asset in snapshot.tables
    ]


def _columns_rows(snapshot: "MetadataSnapshot", tags: _Tags) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for table in snapshot.tables:
        for column in table.columns:
            rows.append(
                [
                    column.semantic_uri,
                    table.name,
                    column.name,
                    ", ".join(column.aliases),
                    column.data_type,
                    tags.asset(column.ref.parts),
                    column.description,
                ]
            )
    return rows


def _relationships_rows(snapshot: "MetadataSnapshot", tags: _Tags) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for edge in snapshot.relationships:
        target = edge.target.parts[-1] if edge.target is not None else ""
        rows.append(
            [
                edge.semantic_uri,
                edge.alias or edge.id,
                edge.source.parts[-1],
                edge.source_column,
                target,
                edge.target_column,
                edge.cardinality,
                edge.kind,
                edge.needs_review,
                edge.version,
                edge.owner.id if edge.owner is not None else "",
                tags.relationship(edge.id),
            ]
        )
    return rows


def _views_rows(org_id: str, tables: list["Table"], tags: _Tags) -> list[list[Any]]:
    return [
        [
            table_uri(org_id, table),
            _business_name(table),
            table.domain_id,
            table.source_id,
            tags.asset(table_ref(table).parts),
            _view_definition(table),
            table.description,
        ]
        for table in tables
        if _is_view(table) and not table.materialize
    ]


def _materialized_rows(org_id: str, tables: list["Table"], tags: _Tags) -> list[list[Any]]:
    return [
        [
            table_uri(org_id, table),
            _business_name(table),
            table.domain_id,
            table.mv_refresh_interval,
            table.mv_consistency,
            table.mv_persist,
            table.mv_incremental,
            tags.asset(table_ref(table).parts),
            _view_definition(table),
        ]
        for table in tables
        if table.materialize and _is_view(table)
    ]


def _metrics_rows(org_id: str, config: "ProvisaConfig", visible: frozenset[str]) -> list[list[Any]]:
    """Metrics whose every referenced table is one the reviewer is being shown.

    ALL, not any: a metric expression names its tables in plain text, so listing one whose base
    table was withheld would print the withheld table's name in the definition column.
    """
    from provisa.compiler.metric_expand import metric_reference_tables

    rows: list[list[Any]] = []
    for metric in config.metrics:
        referenced = set(metric_reference_tables(metric.name, metric.expression))
        if not referenced <= visible:
            continue
        rows.append(
            [
                metric_uri(org_id, metric.name),
                metric.name,
                metric.expression,
                metric.datatype or "",
                metric.description or "",
                metric.ai_context or "",
                metric.from_fact or "",
            ]
        )
    return rows


def _glossary_rows(snapshot: "MetadataSnapshot") -> list[list[Any]]:
    return [
        [
            term.semantic_uri,
            term.name,
            term.definition or "",
            term.is_abstract,
            term.deprecated,
            ", ".join("/".join(ref.parts) for ref in term.refs),
            ", ".join(term.experts),
        ]
        for term in snapshot.glossary_terms
    ]


def sheet_rows(  # REQ-1592
    snapshot: "MetadataSnapshot",
    config: "ProvisaConfig",
    *,
    org_id: str,
) -> dict[str, list[list[Any]]]:
    """Every object sheet's rows, keyed by sheet name, from an ALREADY-NARROWED snapshot.

    Narrowing (rights, domains) is the router's job and has happened before this is called; this
    function's only filtering is the snapshot-to-config alignment described in the module docstring.
    """
    # Aligned on the semantic URI rather than on a bare table name: a name is ambiguous across
    # sources (two sources may both have ``orders``), and the URI is the address the reviewer's
    # id column already carries.
    visible = frozenset(asset.semantic_uri for asset in snapshot.tables)
    tables = [table for table in config.tables if table_uri(org_id, table) in visible]
    # A metric expression names its tables by SEMANTIC name — the alias when there is one, the
    # table name otherwise — so the visibility set the metric filter uses is spelled that way.
    metric_names = frozenset(_business_name(table) for table in tables)
    tags = _tag_index(snapshot)
    return {
        SOURCES: _sources_rows(snapshot, tags),
        TABLES: _tables_rows(snapshot, tags),
        COLUMNS: _columns_rows(snapshot, tags),
        RELATIONSHIPS: _relationships_rows(snapshot, tags),
        VIEWS: _views_rows(org_id, tables, tags),
        METRICS: _metrics_rows(org_id, config, metric_names),
        MATERIALIZED_VIEWS: _materialized_rows(org_id, tables, tags),
        GLOSSARY: _glossary_rows(snapshot),
    }


def _write_report_sheet(worksheet, header: ReportHeader) -> None:
    from openpyxl.styles import Font

    rows = [
        ("Organization", header.org_id),
        ("Environment", header.environment),
        ("Domains covered", header.domains),
        ("Generated", header.generated_at.isoformat()),
        ("Sheets included", ", ".join(header.included)),
        ("Sheets omitted", ", ".join(header.omitted) if header.omitted else "none"),
    ]
    worksheet.append(["field", "value"])
    for row in rows:
        worksheet.append(list(row))
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 20
    worksheet.column_dimensions["B"].width = 80


def _write_object_sheet(worksheet, headers: tuple[str, ...], rows: list[list[Any]]) -> None:
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(row)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"
    # The autofilter spans the header even when there are no rows: a reviewer opening an empty
    # sheet must still see the columns are filterable rather than think the sheet is broken.
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(rows) + 1, 2)}"
    for index in range(1, len(headers) + 1):
        worksheet.column_dimensions[get_column_letter(index)].width = 28


def build_workbook(  # REQ-1592
    header: ReportHeader,
    rows_by_sheet: dict[str, list[list[Any]]],
) -> bytes:
    """The XLSX bytes: the Report sheet first, then each included object sheet in workbook order."""
    from openpyxl import Workbook

    workbook = Workbook()
    report = workbook.active
    assert report is not None  # a fresh Workbook always has one sheet
    report.title = REPORT
    _write_report_sheet(report, header)
    for name in OBJECT_SHEETS:
        if name not in header.included:
            continue
        _write_object_sheet(workbook.create_sheet(name), HEADERS[name], rows_by_sheet[name])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
