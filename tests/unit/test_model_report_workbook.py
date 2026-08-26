# Copyright (c) 2026 Kenneth Stott
# Canary: 4e81b3c6-5d2a-42f7-9b04-8c37fa1d6e02
#
# This source code is licensed under the Business Source License 1.1
# found in the LICENSE file in the root directory of this source tree.

"""REQ-1592: the model as a reviewable workbook.

What is asserted here is what makes the file REVIEWABLE rather than merely produced: a stable
header on every sheet, a frozen header row, an autofilter, an id on every row, the Report sheet
leading and naming what was omitted, and the sheet contents coming from the governed projection
rather than from the raw config.
"""

# Requirements: REQ-1592

from __future__ import annotations

import io
from datetime import UTC, datetime

import openpyxl
import pytest

from provisa.api.metadata_export import workbook as wb
from provisa.api.metadata_export.builder import build_snapshot
from provisa.api.metadata_export.refs import metric_uri
from provisa.core.models import (
    Column,
    Domain,
    Metric,
    ProvisaConfig,
    Source,
    SourceType,
    Table,
    ViewMetricsSpec,
)

ORG = "acme"


def _table(name: str, **kwargs) -> Table:
    columns = kwargs.pop("columns", None) or [
        Column(name="id", data_type="integer", visible_to=["*"])
    ]
    return Table(
        source_id="wh",
        domain_id=kwargs.pop("domain_id", "sales"),
        schema_name="public",
        table_name=name,
        description=kwargs.pop("description", f"{name} table"),
        data_product=kwargs.pop("data_product", True),
        columns=columns,
        **kwargs,
    )


def _config() -> ProvisaConfig:
    return ProvisaConfig(
        sources=[Source(id="wh", type=SourceType.postgresql, description="warehouse")],
        domains=[Domain(id="sales", description="Sales", steward="data-steward")],
        tables=[
            _table(
                "orders",
                modeling_role="fact",
                columns=[
                    Column(name="id", data_type="integer", visible_to=["*"]),
                    Column(name="amount", data_type="numeric", visible_to=["*"]),
                ],
            ),
            _table("staging", data_product=False),
            _table("order_totals", view_sql="SELECT 1", description="a view"),
            _table(
                "order_rollup",
                view_metrics=ViewMetricsSpec(metrics=["revenue"], dimensions=["orders.id"]),
                materialize=True,
                mv_refresh_interval=900,
                description="a materialized view",
            ),
        ],
        metrics=[
            Metric(name="revenue", expression="SUM(orders.amount)", description="gross revenue"),
            Metric(name="held", expression="SUM(staging.id)"),
        ],
        roles=[],
    )


def _rows() -> dict[str, list[list]]:
    config = _config()
    # REQ-1592: the report is a steward's view of the WHOLE registered model, so it projects with
    # the export filter off — exactly as provisa/api/admin/report_router.py does.
    snapshot = build_snapshot(config, org_id=ORG, dialect="postgres", data_products_only=False)
    return wb.sheet_rows(snapshot, config, org_id=ORG)


def _header(**kwargs) -> wb.ReportHeader:
    defaults = dict(
        org_id=ORG,
        environment="prod",
        domains="all",
        generated_at=datetime(2026, 8, 26, 12, 0, tzinfo=UTC),
        included=wb.OBJECT_SHEETS,
        omitted=(),
    )
    defaults.update(kwargs)
    return wb.ReportHeader(**defaults)  # type: ignore[arg-type]


def _load(header: wb.ReportHeader, rows: dict[str, list[list]]):
    return openpyxl.load_workbook(io.BytesIO(wb.build_workbook(header, rows)))


def test_report_sheet_leads_and_names_the_scope():
    book = _load(_header(domains="sales", omitted=(wb.GLOSSARY,), included=(wb.TABLES,)), _rows())
    assert book.sheetnames[0] == wb.REPORT
    stated = {row[0]: row[1] for row in book[wb.REPORT].iter_rows(min_row=2, values_only=True)}
    assert stated["Organization"] == ORG
    assert stated["Environment"] == "prod"
    assert stated["Domains covered"] == "sales"
    assert stated["Generated"] == "2026-08-26T12:00:00+00:00"
    assert stated["Sheets included"] == wb.TABLES
    # A partial workbook says so on its face — otherwise it reads as a complete model.
    assert stated["Sheets omitted"] == wb.GLOSSARY


def test_every_object_sheet_is_frozen_filtered_and_id_led():
    book = _load(_header(), _rows())
    assert book.sheetnames == [wb.REPORT, *wb.OBJECT_SHEETS]
    for name in wb.OBJECT_SHEETS:
        sheet = book[name]
        assert sheet.freeze_panes == "A2", name
        assert sheet.auto_filter.ref is not None, name
        headers = tuple(cell.value for cell in sheet[1])
        assert headers == wb.HEADERS[name], name
        assert headers[0] == "id", name


def test_ids_are_the_semantic_uris_a_comment_traces_back_to():
    rows = _rows()
    assert [row[0] for row in rows[wb.SOURCES]] == [f"provisa://{ORG}/sources/wh"]
    assert f"provisa://{ORG}/sales/tables/orders" in {row[0] for row in rows[wb.TABLES]}
    assert f"provisa://{ORG}/sales/tables/orders#field:amount" in {
        row[0] for row in rows[wb.COLUMNS]
    }
    assert metric_uri(ORG, "revenue") in {row[0] for row in rows[wb.METRICS]}


def test_an_unmarked_table_is_reported_with_the_flag_rather_than_withheld():
    # The Data Product flag is a COLUMN the reviewer filters on, not the report's export filter:
    # an unmarked table is one of the things a model review exists to find.
    rows = _rows()
    flag = wb.HEADERS[wb.TABLES].index("data_product")
    marked = {row[1]: row[flag] for row in rows[wb.TABLES]}
    assert marked["staging"] is False
    assert marked["orders"] is True
    # And the metric that references the unmarked table rides along with it.
    assert {row[1] for row in rows[wb.METRICS]} == {"revenue", "held"}


def test_views_and_materialized_views_are_separate_sheets():
    rows = _rows()
    definition = wb.HEADERS[wb.VIEWS].index("definition")
    assert [row[1] for row in rows[wb.VIEWS]] == ["order_totals"]
    assert [row[definition] for row in rows[wb.VIEWS]] == ["SELECT 1"]
    materialized = rows[wb.MATERIALIZED_VIEWS]
    assert [row[1] for row in materialized] == ["order_rollup"]
    assert materialized[0][wb.HEADERS[wb.MATERIALIZED_VIEWS].index("refresh interval (s)")] == 900
    assert materialized[0][wb.HEADERS[wb.MATERIALIZED_VIEWS].index("definition")] == (
        "metrics: revenue; dimensions: orders.id"
    )


def test_a_taggable_row_carries_one_comma_delimited_tags_cell():
    # Derived tags share the column with assigned ones on purpose: to a reviewer asking which
    # tables are facts, `fact` is the same kind of answer as `pii`, filtered the same way.
    rows = _rows()
    tags = wb.HEADERS[wb.TABLES].index("tags")
    by_name = {row[1]: row[tags] for row in rows[wb.TABLES]}
    assert "fact" in by_name["orders"].split(", ")
    assert by_name["staging"] == ""
    for sheet in (wb.SOURCES, wb.COLUMNS, wb.RELATIONSHIPS, wb.VIEWS, wb.MATERIALIZED_VIEWS):
        assert "tags" in wb.HEADERS[sheet], sheet
    # Neither is taggable — `applies_to` never names them, so a tags column would always be blank.
    for sheet in (wb.METRICS, wb.GLOSSARY):
        assert "tags" not in wb.HEADERS[sheet], sheet


def test_an_omitted_sheet_is_absent_rather_than_empty():
    book = _load(_header(included=(wb.TABLES, wb.COLUMNS), omitted=(wb.SOURCES,)), _rows())
    assert book.sheetnames == [wb.REPORT, wb.TABLES, wb.COLUMNS]


def test_an_empty_sheet_still_carries_its_filterable_header():
    rows = _rows()
    rows[wb.GLOSSARY] = []
    sheet = _load(_header(), rows)[wb.GLOSSARY]
    assert tuple(cell.value for cell in sheet[1]) == wb.HEADERS[wb.GLOSSARY]
    assert sheet.auto_filter.ref == "A1:G2"


def test_there_is_no_reader_the_workbook_is_never_an_edit_path():
    # REQ-1592 one-way-only: the module exposes writers, and nothing that parses a workbook back
    # into configuration. A reader appearing here is the requirement being violated, not a feature.
    assert not [
        name
        for name in dir(wb)
        if not name.startswith("__") and ("read" in name or "parse" in name or "load" in name)
    ]


@pytest.mark.parametrize("sheet", wb.OBJECT_SHEETS)
def test_every_sheet_has_a_declared_header(sheet: str):
    assert sheet in wb.HEADERS
