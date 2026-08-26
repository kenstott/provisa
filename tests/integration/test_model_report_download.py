# Copyright (c) 2026 Kenneth Stott
# Canary: 6b0d92a4-8e51-4c37-a9f2-3d84c1e70b56
#
# This source code is licensed under the Business Source License 1.1
# found in the LICENSE file in the root directory of this source tree.
#
# NOTICE: Use of this software for training artificial intelligence or
# machine learning models is strictly prohibited without explicit written
# permission from the copyright holder.

"""REQ-1592: ``GET /admin/report.xlsx`` end to end, on a real registration plane.

Real PG, the real config loader, the real live-config assembly and the real glossary derivation:
the workbook is asserted to carry what the DB actually holds, because the point of sourcing the
sheets from the metadata-export projection is that a reviewer sees what the platform serves rather
than what a YAML file said at boot.

The two rules under test are the ones a unit test cannot reach: a sheet the caller holds no right
to read is OMITTED rather than the download being refused, and the ``domains`` parameter narrows
the workbook the way it narrows the surfaces the sheets mirror.
"""

# Requirements: REQ-1592

from __future__ import annotations

import io
from pathlib import Path
from types import SimpleNamespace
from typing import cast

import openpyxl
import pytest
import pytest_asyncio
from fastapi import Request

from provisa.api.admin import report_router
from provisa.api.metadata_export import workbook as wb
from provisa.core import domain_policy
from provisa.core.config_loader import load_config, parse_config_dict

pytestmark = [pytest.mark.integration]

SCHEMA_SQL = (Path(__file__).parent.parent.parent / "provisa" / "core" / "schema.sql").read_text()

ORG_ID = "acme"

# Every role below holds full domain reach except where the test is about domains; what varies is
# the RIGHTS, because omission-instead-of-refusal is the rule under test.
_ALL_RIGHTS = [
    "source_registration",
    "table_registration",
    "create_relationship",
    "create_view",
    "glossary_read",
]
_ROLES = {
    "steward": {"capabilities": _ALL_RIGHTS, "domain_access": ["*"]},
    # Holds the Tables surface and nothing else: four sheets must simply not be in the file.
    "table_only": {"capabilities": ["table_registration"], "domain_access": ["*"]},
    "sales_only": {"capabilities": _ALL_RIGHTS, "domain_access": ["sales"]},
}


@pytest_asyncio.fixture(scope="module", loop_scope="session")
async def _init_schema(tenant_db):
    async with tenant_db.acquire() as conn:
        await conn.execute(SCHEMA_SQL)


@pytest_asyncio.fixture(autouse=True)
async def _clean(tenant_db, _init_schema):
    domain_policy.reset()
    async with tenant_db.acquire() as conn:
        await conn.execute(
            """
            TRUNCATE glossary_term_domains, glossary_term_experts, glossary_term_edges,
                     glossary_term_refs, glossary_terms, rls_rules, relationships,
                     relationship_candidates, metrics, table_columns, registered_tables,
                     naming_rules, roles, domains, sources CASCADE
            """
        )
    yield
    domain_policy.reset()


def _config() -> dict:
    """Two domains, a base table in each, a view and a materialized view, and one metric.

    ``staging`` is the control: registered, but not a Data Product, so the export projection drops
    it — and with it the metric whose expression names it, since a metric expression prints its
    base tables in plain text.
    """
    src = {
        "id": "pg1",
        "type": "postgresql",
        "host": "localhost",
        "port": 5432,
        "database": "d",
        "username": "u",
        "password": "p",
    }
    return {
        "sources": [src],
        "domains": [{"id": "sales"}, {"id": "petstore"}],
        "tables": [
            {
                "source_id": "pg1",
                "domain_id": "sales",
                "schema": "public",
                "table": "orders",
                "data_product": True,
                "columns": [
                    {"name": "id", "data_type": "integer", "visible_to": ["*"]},
                    {"name": "amount", "data_type": "numeric", "visible_to": ["*"]},
                ],
            },
            {
                "source_id": "pg1",
                "domain_id": "petstore",
                "schema": "public",
                "table": "pets",
                "data_product": True,
                "columns": [{"name": "order_id", "data_type": "integer", "visible_to": ["*"]}],
            },
            {
                "source_id": "pg1",
                "domain_id": "sales",
                "schema": "public",
                "table": "staging",
                "data_product": False,
                "columns": [{"name": "id", "data_type": "integer", "visible_to": ["*"]}],
            },
            {
                "source_id": "pg1",
                "domain_id": "sales",
                "schema": "public",
                "table": "order_totals",
                "data_product": True,
                "view_sql": "SELECT id FROM orders",
                "columns": [{"name": "id", "data_type": "integer", "visible_to": ["*"]}],
            },
            {
                "source_id": "pg1",
                "domain_id": "petstore",
                "schema": "public",
                "table": "pet_rollup",
                "data_product": True,
                "view_sql": "SELECT order_id FROM pets",
                "materialize": True,
                "mv_refresh_interval": 900,
                "columns": [{"name": "order_id", "data_type": "integer", "visible_to": ["*"]}],
            },
        ],
        "metrics": [
            {"name": "revenue", "expression": "SUM(orders.amount)"},
            {"name": "held", "expression": "SUM(staging.id)"},
        ],
        "roles": [
            {
                "id": rid,
                "capabilities": spec["capabilities"],
                "domain_access": spec["domain_access"],
            }
            for rid, spec in _ROLES.items()
        ],
    }


def _request(role_id: str) -> Request:
    """A caller holding exactly one role, carrying the two attributes the endpoint reads."""
    identity = SimpleNamespace(user_id="u1", roles=[role_id])
    return cast(
        Request,
        SimpleNamespace(state=SimpleNamespace(identity=identity, active_org_id=ORG_ID)),
    )


@pytest_asyncio.fixture
async def served(tenant_db, monkeypatch):
    """The registration plane the report reads, wired to the app state the endpoint resolves."""
    from provisa.api import app as app_module
    from provisa.api.admin import config_export

    from provisa.core.repositories import glossary as glossary_repo

    raw = _config()
    async with tenant_db.acquire() as conn:
        await load_config(parse_config_dict(raw), conn)
        # A derived term publishes only once a curator has defined it (REQ-1387), so the
        # curation step is part of the fixture: without it the Glossary sheet is empty for a
        # reason that has nothing to do with the report.
        for term in await glossary_repo.list_terms(conn):
            await glossary_repo.set_definition(conn, term["id"], f"The {term['name']}.")
    monkeypatch.setattr(app_module.state, "tenant_db", tenant_db, raising=False)
    monkeypatch.setattr(app_module.state, "config", None, raising=False)
    monkeypatch.setattr(
        app_module.state,
        "roles",
        {rid: {"id": rid, **spec} for rid, spec in _ROLES.items()},
        raising=False,
    )
    # The file base build_live_config starts from: the DB-backed sections it rebuilds are the ones
    # this test is about, and pointing at the repo's own config file would make the assertions
    # depend on whatever ships there.
    monkeypatch.setattr(config_export, "read_config", lambda: dict(raw))
    return raw


async def _book(role: str, domains: "list[str] | None" = None):
    response = await report_router.model_report(_request(role), domains=domains)
    assert response.media_type == report_router.XLSX_MEDIA_TYPE
    return response, openpyxl.load_workbook(io.BytesIO(response.body))


def _values(sheet, column: int) -> list:
    return [row[column] for row in sheet.iter_rows(min_row=2, values_only=True)]


@pytest.mark.asyncio(loop_scope="session")
async def test_the_download_is_a_workbook_of_the_governed_model(served):
    response, book = await _book("steward")
    assert response.headers["content-disposition"] == (
        f'attachment; filename="provisa-model-{ORG_ID}.xlsx"'
    )
    assert book.sheetnames == [wb.REPORT, *wb.OBJECT_SHEETS]
    names = set(_values(book[wb.TABLES], 1))
    assert {"orders", "pets"} <= names
    # Registered but not a Data Product: the export projection drops it, so the reviewer never
    # sees it — nor the metric whose expression names it.
    assert "staging" not in names
    assert set(_values(book[wb.METRICS], 1)) == {"revenue"}
    assert _values(book[wb.VIEWS], 1) == ["order_totals"]
    assert _values(book[wb.MATERIALIZED_VIEWS], 1) == ["pet_rollup"]


@pytest.mark.asyncio(loop_scope="session")
async def test_every_sheet_is_reviewable_frozen_filtered_and_id_led(served):
    _, book = await _book("steward")
    for name in wb.OBJECT_SHEETS:
        sheet = book[name]
        assert sheet.freeze_panes == "A2", name
        assert sheet.auto_filter.ref is not None, name
        assert tuple(cell.value for cell in sheet[1]) == wb.HEADERS[name], name
        # A reviewer's comment traces back to the object through this column, so no row may
        # reach the file without one.
        assert all(value for value in _values(sheet, 0)), name


@pytest.mark.asyncio(loop_scope="session")
async def test_a_sheet_the_caller_cannot_read_is_omitted_not_refused(served):
    """NARROWED, NOT REFUSED — the whole download must not 403 over one out-of-reach sheet."""
    _, book = await _book("table_only")
    assert book.sheetnames == [wb.REPORT, wb.TABLES, wb.COLUMNS]
    stated = {row[0]: row[1] for row in book[wb.REPORT].iter_rows(min_row=2, values_only=True)}
    # A partial workbook says on its face what is missing from it.
    for name in (wb.SOURCES, wb.RELATIONSHIPS, wb.VIEWS, wb.METRICS, wb.GLOSSARY):
        assert name in stated["Sheets omitted"]


@pytest.mark.asyncio(loop_scope="session")
async def test_a_domain_scoped_role_receives_its_domain_rather_than_a_403(served):
    _, book = await _book("sales_only")
    assert set(_values(book[wb.TABLES], 1)) == {"orders", "order_totals"}
    assert _values(book[wb.MATERIALIZED_VIEWS], 1) == []
    stated = {row[0]: row[1] for row in book[wb.REPORT].iter_rows(min_row=2, values_only=True)}
    assert stated["Domains covered"] == "sales"
    assert stated["Organization"] == ORG_ID


@pytest.mark.asyncio(loop_scope="session")
async def test_the_domains_parameter_narrows_but_never_widens(served):
    _, narrowed = await _book("steward", domains=["petstore"])
    assert set(_values(narrowed[wb.TABLES], 1)) == {"pets", "pet_rollup"}

    # A scoped role naming a domain it does not hold does not acquire it.
    _, widened = await _book("sales_only", domains=["sales", "petstore"])
    assert set(_values(widened[wb.TABLES], 1)) == {"orders", "order_totals"}


@pytest.mark.asyncio(loop_scope="session")
async def test_narrowing_carries_the_glossary_and_the_source_list_with_it(served):
    """Every sheet follows the tables it addresses — a narrowed workbook leaks no wider shape."""
    _, book = await _book("steward", domains=["petstore"])
    refs = _values(book[wb.GLOSSARY], 5)
    assert refs, "the derivation produced no terms; the narrowing assertion would be vacuous"
    assert not [ref for ref in refs if "orders" in (ref or "")]
    assert set(_values(book[wb.SOURCES], 1)) == {"pg1"}


@pytest.mark.asyncio(loop_scope="session")
async def test_there_is_no_upload_counterpart(served):
    """ONE WAY ONLY: the report route accepts GET and nothing else."""
    methods = {
        (route.path, method)
        for route in report_router.router.routes
        for method in getattr(route, "methods", set())
    }
    assert methods == {("/admin/report.xlsx", "GET")}
